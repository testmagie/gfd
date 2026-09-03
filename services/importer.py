"""
Importer & Exporter for Multi-Sheet Excel and Multi-File CSVs
Supports importing 3+ CSVs simultaneously, multi-tab Excel workbooks,
intelligent header row detection, configurable threshold exclusions,
composite key merge-and-update mechanisms, and conflict resolution strategies.
"""
import os
import io
import csv
import zipfile
import datetime
import re
import uuid
import logging
from typing import Dict, Any, List, Tuple, Optional, Set
import pandas as pd

from services.data_validator import (
    normalize_action_item,
    normalize_decision_item,
    normalize_priority_item,
    sync_companies_and_statuses,
    detect_sheet_type,
    evaluate_row_exclusion,
    get_action_composite_key,
    get_decision_composite_key,
    get_priority_composite_key,
    detect_record_diff,
    parse_date_safely,
    _simplify,
    PALETTE
)

logger = logging.getLogger(__name__)

def _clean_cell_value(val: Any) -> str:
    """Converts any cell value (Timestamp, datetime, float, int, NaN, NaT, None) to a clean, JSON-serializable string."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass

    if isinstance(val, (pd.Timestamp, datetime.datetime, datetime.date)):
        if hasattr(val, 'hour') and val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.strftime('%Y-%m-%d')
        return val.strftime('%Y-%m-%d %H:%M') if hasattr(val, 'strftime') else str(val)

    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)

    s = str(val).strip()
    if s.lower() in ('nan', 'nat', 'none', 'null', '<na>'):
        return ""
    return s

HEADER_KEYWORDS = {
    'item', 'task', 'action', 'status', 'owner', 'company', 'project', 'title',
    'name', 'due', 'date', 'decision', 'priority', 'focus', 'lead', 'summary',
    'deliverable', 'description', 'comments', 'notes', 'timestamp', 'email',
    'function', 'category', 'work', 'issue', 'ticket', 'goal', 'objective',
    'milestone', 'initiative', 'topic', 'group', 'horizon', 'why', 'impact'
}

def detect_excel_header_row(df_raw: pd.DataFrame) -> int:
    """
    Intelligently finds the index of the real header row in an Excel worksheet or table DataFrame.
    Scans the first 15 rows for known table header keywords.
    """
    best_row_idx = 0
    max_score = -1
    num_rows = min(15, len(df_raw))

    for idx in range(num_rows):
        row_vals = [
            str(v).strip().lower()
            for v in df_raw.iloc[idx].values
            if pd.notna(v) and str(v).strip() != '' and str(v).lower() not in ('nan', 'none', 'null')
        ]
        if not row_vals:
            continue

        score = 0
        for val in row_vals:
            val_clean = re.sub(r'[^a-z0-9]', '', val)
            for kw in HEADER_KEYWORDS:
                if kw in val or kw in val_clean:
                    score += 4
                    break
            else:
                if len(val) >= 2 and not val.replace('.', '', 1).isdigit():
                    score += 1

        if score > max_score and (len(row_vals) >= 2 or idx == 0):
            max_score = score
            best_row_idx = idx

    return best_row_idx

def parse_csv_file(file_stream, filename: str, max_rows: Optional[int] = None) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Parses a single CSV file stream into a list of row dicts with intelligent header row detection,
    multi-encoding resilience (UTF-8, UTF-16, Latin-1, CP1252, ISO-8859-1), and delimiter sniffing.
    """
    raw = file_stream.read() if hasattr(file_stream, 'read') else file_stream
    if isinstance(raw, bytes):
        content = None
        # Check for UTF-16 BOM or null bytes
        if raw.startswith(b'\xff\xfe') or raw.startswith(b'\xfe\xff') or (len(raw) > 4 and b'\x00' in raw[:20]):
            for enc in ('utf-16', 'utf-16le', 'utf-16be'):
                try:
                    content = raw.decode(enc)
                    break
                except Exception:
                    continue

        if content is None:
            # Standard single-byte and UTF-8 encodings
            for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'windows-1252', 'iso-8859-1'):
                try:
                    content = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
                except Exception:
                    continue

        if content is None:
            content = raw.decode('latin-1', errors='replace')
    else:
        content = str(raw)

    lines = [l for l in content.splitlines() if l.strip()]
    sheet_name = os.path.splitext(os.path.basename(filename))[0]
    if not lines:
        return sheet_name, []

    # Detect delimiter (comma, tab, semicolon, pipe)
    sample = '\n'.join(lines[:10])
    delimiter = ','
    if '\t' in sample and sample.count('\t') > sample.count(','):
        delimiter = '\t'
    elif ';' in sample and sample.count(';') > sample.count(','):
        delimiter = ';'
    elif '|' in sample and sample.count('|') > sample.count(','):
        delimiter = '|'

    # Find the header row index
    header_idx = 0
    max_score = -1
    for idx, line in enumerate(lines[:15]):
        try:
            reader_row = list(csv.reader([line], delimiter=delimiter))
            if not reader_row or not reader_row[0]:
                continue
            cells = [str(c).strip().lower() for c in reader_row[0] if str(c).strip()]
            if not cells:
                continue
            score = 0
            for c in cells:
                c_clean = re.sub(r'[^a-z0-9]', '', c)
                for kw in HEADER_KEYWORDS:
                    if kw in c or kw in c_clean:
                        score += 4
                        break
                else:
                    if len(c) >= 2 and not c.replace('.', '', 1).isdigit():
                        score += 1
            if score > max_score and (len(cells) >= 2 or idx == 0):
                max_score = score
                header_idx = idx
        except Exception:
            continue

    # Attempt parsing with csv.DictReader first
    records = []
    try:
        f = io.StringIO('\n'.join(lines[header_idx:]))
        reader = csv.DictReader(f, delimiter=delimiter)
        for r in reader:
            cleaned_row = {}
            for k, v in r.items():
                if k is not None:
                    clean_k = str(k).strip()
                    if clean_k and not clean_k.startswith('Unnamed:'):
                        cleaned_row[clean_k] = _clean_cell_value(v)
            if any(v for v in cleaned_row.values()):
                records.append(cleaned_row)
                if max_rows is not None and len(records) > max_rows:
                    raise ValueError(f"CSV exceeds the {max_rows} row limit")
    except Exception as e:
        logger.warning(f"csv.DictReader failed on {filename}: {e}, falling back to pandas parser")
        try:
            f_pandas = io.StringIO('\n'.join(lines[header_idx:]))
            df = pd.read_csv(f_pandas, delimiter=delimiter, on_bad_lines='skip')
            for r in df.to_dict(orient='records'):
                cleaned_row = {
                    str(k).strip(): _clean_cell_value(v)
                    for k, v in r.items()
                    if k is not None and not str(k).startswith('Unnamed:')
                }
                if any(v for v in cleaned_row.values()):
                    records.append(cleaned_row)
        except Exception as e2:
            logger.error(f"Pandas fallback parsing failed on {filename}: {e2}")

    return sheet_name, records

def parse_excel_file(file_stream, max_sheets: Optional[int] = None, max_rows_per_sheet: Optional[int] = None) -> Dict[str, List[Dict[str, Any]]]:
    """
    Parses all worksheets of an uploaded Excel file,
    inspecting sheets with openpyxl, auto-detecting header rows, and cleaning cells.
    """
    if hasattr(file_stream, 'read'):
        content = file_stream.read()
        file_bytes = io.BytesIO(content)
    elif isinstance(file_stream, bytes):
        file_bytes = io.BytesIO(file_stream)
    else:
        file_bytes = file_stream

    excel_file = pd.ExcelFile(file_bytes, engine='openpyxl')
    if max_sheets is not None and len(excel_file.sheet_names) > max_sheets:
        raise ValueError(f"Workbook exceeds the {max_sheets} worksheet limit")
    sheets_data = {}

    for sheet_name in excel_file.sheet_names:
        try:
            # First read raw worksheet without header assumptions
            df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            if max_rows_per_sheet is not None and len(df_raw) > max_rows_per_sheet + 15:
                raise ValueError(f"Worksheet '{sheet_name}' exceeds the {max_rows_per_sheet} row limit")
            if df_raw.empty or df_raw.isna().all().all():
                continue

            # Detect the best header row index
            hdr_idx = detect_excel_header_row(df_raw)
            
            # Read with detected header
            header_vals = df_raw.iloc[hdr_idx].values
            headers = []
            for i, val in enumerate(header_vals):
                clean_h = str(val).strip() if pd.notna(val) else ''
                if not clean_h or clean_h.lower() in ('nan', 'none', 'null') or str(clean_h).startswith('Unnamed:'):
                    clean_h = f"Column_{i+1}"
                headers.append(clean_h)

            df_data = df_raw.iloc[hdr_idx + 1:].copy()
            df_data.columns = headers

            records = []
            for row in df_data.to_dict(orient='records'):
                cleaned_row = {
                    str(k).strip(): _clean_cell_value(v)
                    for k, v in row.items()
                    if k is not None and not str(k).startswith('Column_') and not str(k).startswith('Unnamed:')
                }
                # If standard named columns are empty, check with Column_ keys as fallback
                if not any(cleaned_row.values()):
                    cleaned_row = {
                        str(k).strip(): _clean_cell_value(v)
                        for k, v in row.items()
                        if k is not None
                    }
                if any(cleaned_row.values()):
                    records.append(cleaned_row)

            if records:
                sheets_data[sheet_name] = records
        except Exception as e:
            logger.warning(f"Error parsing worksheet '{sheet_name}': {e}")
            continue

    return sheets_data

def normalize_destination_key(destination: Optional[str]) -> str:
    """Normalizes the import destination parameter."""
    if not destination:
        return "all"
    d = str(destination).strip().lower()
    if d in ["register", "actions", "action", "tasks", "task"]:
        return "register"
    if d in ["decisions", "decision", "dq"]:
        return "decisions"
    if d in ["priorities", "priority", "okr", "focus"]:
        return "priorities"
    if d in ["create_new", "new_company", "company"]:
        return "create_new"
    return "all"

class ImportContext:
    """Manages 1-to-1 matching state for import operations."""
    def __init__(self, current_state: Dict[str, Any]):
        self.actions_list = current_state.get('actions', [])
        self.decisions_list = current_state.get('decisions', [])
        self.priorities_list = current_state.get('priorities', [])

        self.action_id_map: Dict[str, int] = {}
        self.action_spec_map: Dict[str, List[int]] = {}
        self.action_loose_map: Dict[str, List[int]] = {}

        for idx, a in enumerate(self.actions_list):
            if a.get('id'):
                self.action_id_map[str(a['id']).lower()] = idx
            comp = str(a.get('company', '')).strip().lower()
            item = _simplify(a.get('item', ''))
            func = _simplify(a.get('function', ''))
            owner = _simplify(a.get('owner', ''))
            due = _simplify(a.get('due', ''))

            spec_k = f"{comp}::{func}::{owner}::{item}::{due}"
            self.action_spec_map.setdefault(spec_k, []).append(idx)
            loose_k = f"{comp}::{item}"
            self.action_loose_map.setdefault(loose_k, []).append(idx)

        self.matched_actions: Set[int] = set()
        self.new_actions: List[Dict[str, Any]] = []

        self.decision_id_map: Dict[str, int] = {}
        self.decision_spec_map: Dict[str, List[int]] = {}
        self.decision_loose_map: Dict[str, List[int]] = {}

        for idx, d in enumerate(self.decisions_list):
            if d.get('id'):
                self.decision_id_map[str(d['id']).lower()] = idx
            dec = _simplify(d.get('decision', ''))
            owner = _simplify(d.get('owner', ''))
            self.decision_spec_map.setdefault(f"{dec}::{owner}", []).append(idx)
            self.decision_loose_map.setdefault(dec, []).append(idx)

        self.matched_decisions: Set[int] = set()
        self.new_decisions: List[Dict[str, Any]] = []

        self.priority_id_map: Dict[str, int] = {}
        self.priority_spec_map: Dict[str, List[int]] = {}
        self.priority_loose_map: Dict[str, List[int]] = {}

        for idx, p in enumerate(self.priorities_list):
            if p.get('id'):
                self.priority_id_map[str(p['id']).lower()] = idx
            group = str(p.get('group', '')).strip().lower()
            focus = _simplify(p.get('focusArea', ''))
            self.priority_spec_map.setdefault(f"{group}::{focus}", []).append(idx)
            self.priority_loose_map.setdefault(focus, []).append(idx)

        self.matched_priorities: Set[int] = set()
        self.new_priorities: List[Dict[str, Any]] = []

def _process_action_record(norm: Dict[str, Any], ctx: ImportContext, mode_key: str, strategy_key: str, metrics: Dict[str, Any], sheet_stat: Dict[str, Any]):
    """Processes a single normalized action record through 1-to-1 matching and conflict resolution."""
    if mode_key == "replace":
        norm["id"] = norm.get("id") or f"a_{uuid.uuid4().hex[:8]}"
        ctx.new_actions.append(norm)
        metrics["appended"] += 1
        metrics["actions"] += 1
        sheet_stat["appended"] += 1
        return

    if mode_key == "append":
        norm["id"] = f"a_{uuid.uuid4().hex[:8]}"
        ctx.new_actions.append(norm)
        metrics["appended"] += 1
        metrics["actions"] += 1
        sheet_stat["appended"] += 1
        return

    # Merge & Update Mode: 1-to-1 matching against existing records only
    matched_idx = None
    if norm.get('id') and str(norm['id']).lower() in ctx.action_id_map:
        cand = ctx.action_id_map[str(norm['id']).lower()]
        if cand not in ctx.matched_actions:
            matched_idx = cand

    comp = str(norm.get('company', '')).strip().lower()
    item = _simplify(norm.get('item', ''))
    func = _simplify(norm.get('function', ''))
    owner = _simplify(norm.get('owner', ''))
    due = _simplify(norm.get('due', ''))

    if matched_idx is None:
        spec_k = f"{comp}::{func}::{owner}::{item}::{due}"
        for cand in ctx.action_spec_map.get(spec_k, []):
            if cand not in ctx.matched_actions:
                matched_idx = cand
                break

    if matched_idx is None:
        loose_k = f"{comp}::{item}"
        for cand in ctx.action_loose_map.get(loose_k, []):
            if cand not in ctx.matched_actions:
                matched_idx = cand
                break

    if matched_idx is not None:
        ctx.matched_actions.add(matched_idx)
        existing_act = ctx.actions_list[matched_idx]
        diffs = detect_record_diff(existing_act, norm, ["function", "status", "owner", "founderDependency", "due", "comments"])

        if not diffs:
            metrics["merged"] += 1
            metrics["actions"] += 1
            return

        if strategy_key == "manual_review":
            metrics["flagged"] += 1
            metrics["actions"] += 1
            sheet_stat["flagged"] += 1
            metrics["conflicts"].append({
                "type": "action",
                "id": existing_act["id"],
                "existing": existing_act,
                "incoming": norm,
                "diffs": {k: {"existing": v[0], "incoming": v[1]} for k, v in diffs.items()}
            })
            return

        # Field-level update
        metrics["updated"] += 1
        metrics["actions"] += 1
        sheet_stat["updated"] += 1

        if strategy_key == "existing_wins":
            for k, (e_val, i_val) in diffs.items():
                if not e_val and i_val:
                    existing_act[k] = i_val
        elif strategy_key == "timestamp_wins":
            e_date = parse_date_safely(existing_act.get('due'))
            i_date = parse_date_safely(norm.get('due'))
            if i_date and (not e_date or i_date >= e_date):
                for k, (_, i_val) in diffs.items():
                    if i_val:
                        existing_act[k] = i_val
        else:  # incoming_wins (default)
            for k, (_, i_val) in diffs.items():
                if i_val:
                    existing_act[k] = i_val
    else:
        norm["id"] = norm.get("id") or f"a_{uuid.uuid4().hex[:8]}"
        ctx.new_actions.append(norm)
        metrics["appended"] += 1
        metrics["actions"] += 1
        sheet_stat["appended"] += 1

def _process_decision_record(norm: Dict[str, Any], ctx: ImportContext, mode_key: str, strategy_key: str, metrics: Dict[str, Any], sheet_stat: Dict[str, Any]):
    """Processes a single normalized decision record through 1-to-1 matching and conflict resolution."""
    if mode_key == "replace":
        norm["id"] = norm.get("id") or f"d_{uuid.uuid4().hex[:8]}"
        ctx.new_decisions.append(norm)
        metrics["appended"] += 1
        metrics["decisions"] += 1
        sheet_stat["appended"] += 1
        return

    if mode_key == "append":
        norm["id"] = f"d_{uuid.uuid4().hex[:8]}"
        ctx.new_decisions.append(norm)
        metrics["appended"] += 1
        metrics["decisions"] += 1
        sheet_stat["appended"] += 1
        return

    matched_idx = None
    if norm.get('id') and str(norm['id']).lower() in ctx.decision_id_map:
        cand = ctx.decision_id_map[str(norm['id']).lower()]
        if cand not in ctx.matched_decisions:
            matched_idx = cand

    dec = _simplify(norm.get('decision', ''))
    owner = _simplify(norm.get('owner', ''))

    if matched_idx is None:
        spec_k = f"{dec}::{owner}"
        for cand in ctx.decision_spec_map.get(spec_k, []):
            if cand not in ctx.matched_decisions:
                matched_idx = cand
                break

    if matched_idx is None:
        for cand in ctx.decision_loose_map.get(dec, []):
            if cand not in ctx.matched_decisions:
                matched_idx = cand
                break

    if matched_idx is not None:
        ctx.matched_decisions.add(matched_idx)
        existing_dec = ctx.decisions_list[matched_idx]
        diffs = detect_record_diff(existing_dec, norm, ["owner", "status", "founderDependency", "impact", "deadline", "nextReview"])

        if not diffs:
            metrics["merged"] += 1
            metrics["decisions"] += 1
            return

        if strategy_key == "manual_review":
            metrics["flagged"] += 1
            metrics["decisions"] += 1
            sheet_stat["flagged"] += 1
            metrics["conflicts"].append({
                "type": "decision",
                "id": existing_dec["id"],
                "existing": existing_dec,
                "incoming": norm,
                "diffs": {k: {"existing": v[0], "incoming": v[1]} for k, v in diffs.items()}
            })
            return

        metrics["updated"] += 1
        metrics["decisions"] += 1
        sheet_stat["updated"] += 1

        if strategy_key == "existing_wins":
            for k, (e_val, i_val) in diffs.items():
                if not e_val and i_val:
                    existing_dec[k] = i_val
        elif strategy_key == "timestamp_wins":
            e_date = parse_date_safely(existing_dec.get('deadline') or existing_dec.get('nextReview'))
            i_date = parse_date_safely(norm.get('deadline') or norm.get('nextReview'))
            if i_date and (not e_date or i_date >= e_date):
                for k, (_, i_val) in diffs.items():
                    if i_val:
                        existing_dec[k] = i_val
        else:  # incoming_wins
            for k, (_, i_val) in diffs.items():
                if i_val:
                    existing_dec[k] = i_val
    else:
        norm["id"] = norm.get("id") or f"d_{uuid.uuid4().hex[:8]}"
        ctx.new_decisions.append(norm)
        metrics["appended"] += 1
        metrics["decisions"] += 1
        sheet_stat["appended"] += 1

def _process_priority_record(norm: Dict[str, Any], ctx: ImportContext, mode_key: str, strategy_key: str, metrics: Dict[str, Any], sheet_stat: Dict[str, Any]):
    """Processes a single normalized priority record through 1-to-1 matching and conflict resolution."""
    if mode_key == "replace":
        norm["id"] = norm.get("id") or f"p_{uuid.uuid4().hex[:8]}"
        ctx.new_priorities.append(norm)
        metrics["appended"] += 1
        metrics["priorities"] += 1
        sheet_stat["appended"] += 1
        return

    if mode_key == "append":
        norm["id"] = f"p_{uuid.uuid4().hex[:8]}"
        ctx.new_priorities.append(norm)
        metrics["appended"] += 1
        metrics["priorities"] += 1
        sheet_stat["appended"] += 1
        return

    matched_idx = None
    if norm.get('id') and str(norm['id']).lower() in ctx.priority_id_map:
        cand = ctx.priority_id_map[str(norm['id']).lower()]
        if cand not in ctx.matched_priorities:
            matched_idx = cand

    group = str(norm.get('group', '')).strip().lower()
    focus = _simplify(norm.get('focusArea', ''))

    if matched_idx is None:
        spec_k = f"{group}::{focus}"
        for cand in ctx.priority_spec_map.get(spec_k, []):
            if cand not in ctx.matched_priorities:
                matched_idx = cand
                break

    if matched_idx is None:
        for cand in ctx.priority_loose_map.get(focus, []):
            if cand not in ctx.matched_priorities:
                matched_idx = cand
                break

    if matched_idx is not None:
        ctx.matched_priorities.add(matched_idx)
        existing_prio = ctx.priorities_list[matched_idx]
        diffs = detect_record_diff(existing_prio, norm, ["group", "priority", "why", "horizon"])

        if not diffs:
            metrics["merged"] += 1
            metrics["priorities"] += 1
            return

        if strategy_key == "manual_review":
            metrics["flagged"] += 1
            metrics["priorities"] += 1
            sheet_stat["flagged"] += 1
            metrics["conflicts"].append({
                "type": "priority",
                "id": existing_prio["id"],
                "existing": existing_prio,
                "incoming": norm,
                "diffs": {k: {"existing": v[0], "incoming": v[1]} for k, v in diffs.items()}
            })
            return

        metrics["updated"] += 1
        metrics["priorities"] += 1
        sheet_stat["updated"] += 1

        if strategy_key == "existing_wins":
            for k, (e_val, i_val) in diffs.items():
                if not e_val and i_val:
                    existing_prio[k] = i_val
        elif strategy_key == "timestamp_wins":
            for k, (_, i_val) in diffs.items():
                if i_val:
                    existing_prio[k] = i_val
        else:  # incoming_wins
            for k, (_, i_val) in diffs.items():
                if i_val:
                    existing_prio[k] = i_val
    else:
        norm["id"] = norm.get("id") or f"p_{uuid.uuid4().hex[:8]}"
        ctx.new_priorities.append(norm)
        metrics["appended"] += 1
        metrics["priorities"] += 1
        sheet_stat["appended"] += 1

def process_dataset_import(
    sheets_data: Dict[str, List[Dict[str, Any]]],
    current_state: Dict[str, Any],
    destination: str = "all",
    mode: str = "merge",
    conflict_strategy: str = "incoming_wins",
    min_quality_score: float = 0.0,
    excluded_statuses: Optional[Set[str]] = None,
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    new_company_name: Optional[str] = None
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Core dataset import engine supporting:
    - Multi-tab Excel and 3+ CSV files
    - Configurable threshold exclusions (date range, quality score, excluded statuses)
    - Overlap and duplicate detection via composite keys
    - Field-level merge and update with configurable conflict strategies:
      * 'incoming_wins' (Default): Incoming non-empty fields update existing records
      * 'existing_wins': Preserves existing values; incoming only fills empty fields
      * 'timestamp_wins': Newer date/due wins
      * 'manual_review': Flags conflicting records for manual review queue
    - Dynamic Import Destinations: 'all', 'register', 'decisions', 'priorities', 'create_new'
    - Dynamic Import Strategies: 'merge', 'append', 'replace', 'delete_merge'
      * 'delete_merge': Like 'merge' but also removes dashboard records absent from the CSV
    - Detailed feedback metrics (appended, updated, skipped, flagged, deleted, per-sheet breakdown)
    """
    dest_key = normalize_destination_key(destination)
    mode_key = str(mode).strip().lower() if mode else "merge"
    strategy_key = str(conflict_strategy).strip().lower() if conflict_strategy else "incoming_wins"

    # delete_merge always preserves dashboard edits: only fill blank dashboard fields
    # from the CSV — never overwrite a value the user has already set in the dashboard.
    if mode_key == "delete_merge":
        strategy_key = "existing_wins"

    # Tracking counters and metadata
    metrics = {
        "appended": 0,
        "updated": 0,
        "merged": 0,
        "skipped": 0,
        "flagged": 0,
        "deleted": 0,
        "sheets_processed": 0,
        "actions": 0,
        "decisions": 0,
        "priorities": 0,
        "sheet_names": list(sheets_data.keys()),
        "exclusion_reasons": {},
        "sheet_breakdown": {},
        "conflicts": [],
        "total_processed": 0,
        "actions_total": 0,
        "decisions_total": 0,
        "priorities_total": 0
    }

    # Handle 'create_new' company registration
    target_company = None
    if dest_key == "create_new":
        target_company = (new_company_name or "").strip()
        if not target_company and sheets_data:
            # Infer from first sheet name
            first_sheet = list(sheets_data.keys())[0]
            if first_sheet.lower() not in ('sheet1', 'data', 'actions', 'export', 'upload'):
                target_company = first_sheet.strip()
            else:
                target_company = f"Company_{datetime.datetime.now().strftime('%m%d')}"
        elif not target_company:
            target_company = "New Company"

        settings = current_state.setdefault('settings', {})
        companies = settings.setdefault('companies', [])
        company_colors = settings.setdefault('companyColors', {})
        if not any(c.get('id', '').lower() == target_company.lower() for c in companies):
            companies.append({'id': target_company, 'name': target_company})
            color_idx = len(companies) % len(PALETTE)
            company_colors[target_company] = PALETTE[color_idx]

    ctx = ImportContext(current_state)

    for sheet_name, rows in sheets_data.items():
        if not rows:
            continue
        metrics["sheets_processed"] += 1
        name_lower = str(sheet_name).lower().strip()

        sheet_stat = {"total": len(rows), "appended": 0, "updated": 0, "skipped": 0, "flagged": 0}
        metrics["total_processed"] += len(rows)

        # Determine target category for rows in this sheet
        if dest_key == "create_new":
            # Map all rows as actions for target_company
            for r in rows:
                norm = normalize_action_item(r, default_company=target_company)
                if not norm:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                    continue

                is_excl, reason = evaluate_row_exclusion(
                    norm, record_type='actions', min_quality_score=min_quality_score,
                    excluded_statuses=excluded_statuses, date_start=date_start, date_end=date_end
                )
                if is_excl:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    r_key = reason.split(':')[0] if reason else 'excluded'
                    metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                    continue

                norm['company'] = target_company
                _process_action_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

        elif dest_key == "register":
            default_comp = "General"
            if not any(k == name_lower for k in ['actions', 'action items', 'tasks', 'register', 'sheet1', 'data', 'general', 'upload']):
                default_comp = str(sheet_name).strip()
            for r in rows:
                norm = normalize_action_item(r, default_company=default_comp)
                if not norm:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                    continue

                is_excl, reason = evaluate_row_exclusion(
                    norm, record_type='actions', min_quality_score=min_quality_score,
                    excluded_statuses=excluded_statuses, date_start=date_start, date_end=date_end
                )
                if is_excl:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    r_key = reason.split(':')[0] if reason else 'excluded'
                    metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                    continue

                _process_action_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

        elif dest_key == "decisions":
            for r in rows:
                norm = normalize_decision_item(r)
                if not norm:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                    continue

                is_excl, reason = evaluate_row_exclusion(
                    norm, record_type='decisions', min_quality_score=min_quality_score,
                    excluded_statuses=excluded_statuses, date_start=date_start, date_end=date_end
                )
                if is_excl:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    r_key = reason.split(':')[0] if reason else 'excluded'
                    metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                    continue

                _process_decision_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

        elif dest_key == "priorities":
            for r in rows:
                norm = normalize_priority_item(r)
                if not norm:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                    continue

                is_excl, reason = evaluate_row_exclusion(
                    norm, record_type='priorities', min_quality_score=min_quality_score,
                    excluded_statuses=excluded_statuses, date_start=date_start, date_end=date_end
                )
                if is_excl:
                    metrics["skipped"] += 1
                    sheet_stat["skipped"] += 1
                    r_key = reason.split(':')[0] if reason else 'excluded'
                    metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                    continue

                _process_priority_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

        else:
            # dest_key == 'all': Smart auto-detection with fallback cascade
            detected_type = detect_sheet_type(sheet_name, rows)
            
            if detected_type == 'decisions':
                for r in rows:
                    norm = normalize_decision_item(r)
                    if not norm:
                        # Fallback try action item
                        norm_act = normalize_action_item(r, default_company="General")
                        if norm_act:
                            is_excl, reason = evaluate_row_exclusion(norm_act, 'actions', min_quality_score, excluded_statuses, date_start, date_end)
                            if is_excl:
                                metrics["skipped"] += 1
                                sheet_stat["skipped"] += 1
                                continue
                            _process_action_record(norm_act, ctx, mode_key, strategy_key, metrics, sheet_stat)
                        else:
                            metrics["skipped"] += 1
                            sheet_stat["skipped"] += 1
                            metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                        continue

                    is_excl, reason = evaluate_row_exclusion(norm, 'decisions', min_quality_score, excluded_statuses, date_start, date_end)
                    if is_excl:
                        metrics["skipped"] += 1
                        sheet_stat["skipped"] += 1
                        r_key = reason.split(':')[0] if reason else 'excluded'
                        metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                        continue

                    _process_decision_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

            elif detected_type == 'priorities':
                for r in rows:
                    norm = normalize_priority_item(r)
                    if not norm:
                        norm_act = normalize_action_item(r, default_company="General")
                        if norm_act:
                            is_excl, reason = evaluate_row_exclusion(norm_act, 'actions', min_quality_score, excluded_statuses, date_start, date_end)
                            if is_excl:
                                metrics["skipped"] += 1
                                sheet_stat["skipped"] += 1
                                continue
                            _process_action_record(norm_act, ctx, mode_key, strategy_key, metrics, sheet_stat)
                        else:
                            metrics["skipped"] += 1
                            sheet_stat["skipped"] += 1
                            metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                        continue

                    is_excl, reason = evaluate_row_exclusion(norm, 'priorities', min_quality_score, excluded_statuses, date_start, date_end)
                    if is_excl:
                        metrics["skipped"] += 1
                        sheet_stat["skipped"] += 1
                        r_key = reason.split(':')[0] if reason else 'excluded'
                        metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                        continue

                    _process_priority_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

            else:
                # Default: Actions or company-specific sheet
                default_comp = "General"
                if not any(k == name_lower for k in ['actions', 'action items', 'tasks', 'register', 'sheet1', 'data', 'general', 'upload']):
                    default_comp = str(sheet_name).strip()
                for r in rows:
                    norm = normalize_action_item(r, default_company=default_comp)
                    if not norm:
                        # Fallback try decision then priority
                        norm_dec = normalize_decision_item(r)
                        if norm_dec:
                            _process_decision_record(norm_dec, ctx, mode_key, strategy_key, metrics, sheet_stat)
                            continue
                        norm_prio = normalize_priority_item(r)
                        if norm_prio:
                            _process_priority_record(norm_prio, ctx, mode_key, strategy_key, metrics, sheet_stat)
                            continue
                        metrics["skipped"] += 1
                        sheet_stat["skipped"] += 1
                        metrics["exclusion_reasons"]["empty_or_invalid"] = metrics["exclusion_reasons"].get("empty_or_invalid", 0) + 1
                        continue

                    is_excl, reason = evaluate_row_exclusion(norm, 'actions', min_quality_score, excluded_statuses, date_start, date_end)
                    if is_excl:
                        metrics["skipped"] += 1
                        sheet_stat["skipped"] += 1
                        r_key = reason.split(':')[0] if reason else 'excluded'
                        metrics["exclusion_reasons"][r_key] = metrics["exclusion_reasons"].get(r_key, 0) + 1
                        continue

                    _process_action_record(norm, ctx, mode_key, strategy_key, metrics, sheet_stat)

        metrics["sheet_breakdown"][sheet_name] = sheet_stat

    # Apply results based on mode
    if mode_key == "replace":
        if dest_key == "register":
            current_state["actions"] = ctx.new_actions
        elif dest_key == "decisions":
            current_state["decisions"] = ctx.new_decisions
        elif dest_key == "priorities":
            current_state["priorities"] = ctx.new_priorities
        elif dest_key == "create_new":
            filtered_actions = [a for a in current_state.get("actions", []) if a.get("company") != target_company]
            filtered_actions.extend(ctx.new_actions)
            current_state["actions"] = filtered_actions
        else:
            if ctx.new_actions:
                current_state["actions"] = ctx.new_actions
            if ctx.new_decisions:
                current_state["decisions"] = ctx.new_decisions
            if ctx.new_priorities:
                current_state["priorities"] = ctx.new_priorities
    elif mode_key == "delete_merge":
        # Merge + delete unmatched: Remove any existing record that was NOT matched
        # by an incoming CSV row. Dashboard edits on MATCHED records are preserved.
        if dest_key in ("register", "all", "create_new"):
            kept_actions = [
                a for idx, a in enumerate(ctx.actions_list)
                if idx in ctx.matched_actions
            ]
            metrics["deleted"] += len(ctx.actions_list) - len(kept_actions)
            current_state["actions"] = kept_actions + ctx.new_actions
        if dest_key in ("decisions", "all"):
            kept_decisions = [
                d for idx, d in enumerate(ctx.decisions_list)
                if idx in ctx.matched_decisions
            ]
            metrics["deleted"] += len(ctx.decisions_list) - len(kept_decisions)
            current_state["decisions"] = kept_decisions + ctx.new_decisions
        if dest_key in ("priorities", "all"):
            kept_priorities = [
                p for idx, p in enumerate(ctx.priorities_list)
                if idx in ctx.matched_priorities
            ]
            metrics["deleted"] += len(ctx.priorities_list) - len(kept_priorities)
            current_state["priorities"] = kept_priorities + ctx.new_priorities
        if dest_key == "create_new":
            # For create_new, only prune actions belonging to the target company
            other_actions = [a for a in current_state.get("actions", []) if a.get("company") != target_company]
            kept_company_actions = [
                a for idx, a in enumerate(ctx.actions_list)
                if idx in ctx.matched_actions and a.get("company") == target_company
            ]
            current_state["actions"] = other_actions + kept_company_actions + ctx.new_actions
    else:
        if ctx.new_actions:
            current_state["actions"] = current_state.get("actions", []) + ctx.new_actions
        if ctx.new_decisions:
            current_state["decisions"] = current_state.get("decisions", []) + ctx.new_decisions
        if ctx.new_priorities:
            current_state["priorities"] = current_state.get("priorities", []) + ctx.new_priorities

    # Ensure companies and statuses are up to date
    current_state = sync_companies_and_statuses(current_state)

    metrics["actions_total"] = len(current_state.get("actions", []))
    metrics["decisions_total"] = len(current_state.get("decisions", []))
    metrics["priorities_total"] = len(current_state.get("priorities", []))

    return current_state, metrics

def export_state_to_excel(state: Dict[str, Any]) -> io.BytesIO:
    """
    Exports the entire dashboard state to a multi-tab, executive-styled Excel workbook (.xlsx).
    Features:
    - Slate/Navy header styling with bold white text
    - Status column color-coded badge styling matching the dashboard palette (Done=Green, WIP=Blue, On Hold=Amber, Future=Purple, Blocked=Red)
    - Excel DataValidation Dropdown lists for Status, Company, Function, Dependency, and Horizon columns
    - Auto-adjusted column widths so no header or text is truncated
    - Auto-filter enabled on all table headers
    - Frozen header pane for seamless scrolling
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Keep reference to remove default sheet later
    default_sheet = wb.active

    # Styling definitions
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10.5, color="0F172A")
    id_font = Font(name="Consolas", size=9.5, color="64748B")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    header_border = Border(
        left=Side(style='thin', color='334155'),
        right=Side(style='thin', color='334155'),
        top=Side(style='medium', color='334155'),
        bottom=Side(style='medium', color='334155')
    )

    # Status color palette mapping (Fill, TextColor)
    status_palette = {
        'done': ('E6F4EA', '0D652D'),
        'completed': ('E6F4EA', '0D652D'),
        'resolved': ('E6F4EA', '0D652D'),
        'closed': ('E6F4EA', '0D652D'),
        'decided': ('E6F4EA', '0D652D'),
        'wip': ('E8F0FE', '185ABC'),
        'in progress': ('E8F0FE', '185ABC'),
        'active': ('E8F0FE', '185ABC'),
        'in review': ('E8F0FE', '185ABC'),
        'to review': ('E8F0FE', '185ABC'),
        'on hold': ('FEF7E0', 'B06000'),
        'pending': ('FEF7E0', 'B06000'),
        'paused': ('FEF7E0', 'B06000'),
        'waiting': ('FEF7E0', 'B06000'),
        'future': ('F3E8FF', '6B21A8'),
        'backlog': ('F3E8FF', '6B21A8'),
        'to start': ('F3E8FF', '6B21A8'),
        'planned': ('F3E8FF', '6B21A8'),
        'blocked': ('FCE8E6', 'C5221F'),
        'attention': ('FCE8E6', 'C5221F'),
        'urgent': ('FCE8E6', 'C5221F'),
        'issue': ('FCE8E6', 'C5221F'),
        'cancelled': ('F1F5F9', '64748B'),
    }

    def get_status_style(status_text: str):
        st_clean = str(status_text).strip().lower()
        if st_clean in status_palette:
            bg_hex, fg_hex = status_palette[st_clean]
            return PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid"), Font(name="Calibri", size=10, bold=True, color=fg_hex)
        for k, (bg_hex, fg_hex) in status_palette.items():
            if k in st_clean:
                return PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid"), Font(name="Calibri", size=10, bold=True, color=fg_hex)
        return PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), Font(name="Calibri", size=10, bold=True, color="334155")

    # Helper: Populate Lookups Sheet for dynamic dropdown lists
    ws_lookups = wb.create_sheet(title="_Lookups")
    ws_lookups.views.sheetView[0].tabSelected = False
    ws_lookups.sheet_state = 'hidden'

    # Extract distinct lookups
    all_companies = [c.get('name', c.get('id', '')) for c in state.get('settings', {}).get('companies', []) if c.get('name') or c.get('id')]
    actions_companies = [str(a.get('company', '')).strip() for a in state.get('actions', []) if str(a.get('company', '')).strip()]
    unique_companies = sorted(list(set(all_companies + actions_companies))) or ["General", "Pranik", "Abhi", "Aarna", "Miraa", "EdT"]
    
    unique_statuses = [
        "Done", "In Progress", "WIP", "On Hold", "Future", "Blocked", "Pending", "To Review"
    ]
    unique_functions = [
        "GTM", "Product", "Engineering", "Operations", "Finance", "People", "Content", "Marketing", "Sales", "Legal", "General"
    ]
    unique_dependencies = [
        "Decision", "Clarity", "Blocker", "To Review", "None"
    ]
    unique_horizons = [
        "Next 30 days", "Next 60 days", "Next 90 days", "Q1", "Q2", "Q3", "Q4", "Long-term"
    ]

    # Write lookups columns
    ws_lookups.append(["Companies", "Statuses", "Functions", "Dependencies", "Horizons"])
    max_len = max(len(unique_companies), len(unique_statuses), len(unique_functions), len(unique_dependencies), len(unique_horizons))
    for i in range(max_len):
        ws_lookups.append([
            unique_companies[i] if i < len(unique_companies) else "",
            unique_statuses[i] if i < len(unique_statuses) else "",
            unique_functions[i] if i < len(unique_functions) else "",
            unique_dependencies[i] if i < len(unique_dependencies) else "",
            unique_horizons[i] if i < len(unique_horizons) else "",
        ])

    comp_range = f"_Lookups!$A$2:$A${len(unique_companies)+1}"
    status_range = f"_Lookups!$B$2:$B${len(unique_statuses)+1}"
    func_range = f"_Lookups!$C$2:$C${len(unique_functions)+1}"
    dep_range = f"_Lookups!$D$2:$D${len(unique_dependencies)+1}"
    horizon_range = f"_Lookups!$E$2:$E${len(unique_horizons)+1}"

    # -------------------------------------------------------------
    # 1. ACTIONS SHEET (Action Items Register)
    # -------------------------------------------------------------
    ws_actions = wb.create_sheet(title="Actions")
    action_labels = ["ID", "Company", "Function", "Action Item", "Status", "Owner", "Founder Dependency", "Due Date", "Comments / Notes"]
    ws_actions.append(action_labels)
    ws_actions.row_dimensions[1].height = 26

    actions_data = state.get('actions', [])
    for row_idx, a in enumerate(actions_data, start=2):
        row_vals = [
            str(a.get('id', f"a_{row_idx-1}")),
            str(a.get('company', '')),
            str(a.get('function', '')),
            str(a.get('item', '')),
            str(a.get('status', 'WIP')),
            str(a.get('owner', '')),
            str(a.get('founderDependency', 'None')),
            str(a.get('due', '')),
            str(a.get('comments', ''))
        ]
        ws_actions.append(row_vals)
        ws_actions.row_dimensions[row_idx].height = 20

    # Data Validations on Actions sheet
    max_act_row = max(len(actions_data)+100, 500)
    dv_act_status = DataValidation(type='list', formula1=status_range, allow_blank=True)
    ws_actions.add_data_validation(dv_act_status)
    dv_act_status.add(f"E2:E{max_act_row}")

    dv_act_comp = DataValidation(type='list', formula1=comp_range, allow_blank=True)
    ws_actions.add_data_validation(dv_act_comp)
    dv_act_comp.add(f"B2:B{max_act_row}")

    dv_act_func = DataValidation(type='list', formula1=func_range, allow_blank=True)
    ws_actions.add_data_validation(dv_act_func)
    dv_act_func.add(f"C2:C{max_act_row}")

    dv_act_dep = DataValidation(type='list', formula1=dep_range, allow_blank=True)
    ws_actions.add_data_validation(dv_act_dep)
    dv_act_dep.add(f"G2:G{max_act_row}")

    # -------------------------------------------------------------
    # 2. DECISIONS SHEET (Decisions Log)
    # -------------------------------------------------------------
    ws_decisions = wb.create_sheet(title="Decisions")
    decision_labels = ["ID", "Decision Required", "Owner", "Status", "Founder Dependency", "Impact if Delayed", "Deadline", "Next Review"]
    ws_decisions.append(decision_labels)
    ws_decisions.row_dimensions[1].height = 26

    decisions_data = state.get('decisions', [])
    for row_idx, d in enumerate(decisions_data, start=2):
        row_vals = [
            str(d.get('id', f"d_{row_idx-1}")),
            str(d.get('decision', '')),
            str(d.get('owner', '')),
            str(d.get('status', 'To Start')),
            str(d.get('founderDependency', 'To Review')),
            str(d.get('impact', '')),
            str(d.get('deadline', '')),
            str(d.get('nextReview', ''))
        ]
        ws_decisions.append(row_vals)
        ws_decisions.row_dimensions[row_idx].height = 20

    max_dec_row = max(len(decisions_data)+100, 500)
    dv_dec_status = DataValidation(type='list', formula1=status_range, allow_blank=True)
    ws_decisions.add_data_validation(dv_dec_status)
    dv_dec_status.add(f"D2:D{max_dec_row}")

    dv_dec_dep = DataValidation(type='list', formula1=dep_range, allow_blank=True)
    ws_decisions.add_data_validation(dv_dec_dep)
    dv_dec_dep.add(f"E2:E{max_dec_row}")

    # -------------------------------------------------------------
    # 3. PRIORITIES SHEET (Strategic Priorities)
    # -------------------------------------------------------------
    ws_priorities = wb.create_sheet(title="Priorities")
    priority_labels = ["ID", "Priority Rank", "Strategic Group", "Focus Area / Initiative", "Why & Strategic Impact", "Time Horizon"]
    ws_priorities.append(priority_labels)
    ws_priorities.row_dimensions[1].height = 26

    priorities_data = state.get('priorities', [])
    for row_idx, p in enumerate(priorities_data, start=2):
        row_vals = [
            str(p.get('id', f"p_{row_idx-1}")),
            str(p.get('priority', '1.0')),
            str(p.get('group', 'Strategic Focus')),
            str(p.get('focusArea', '')),
            str(p.get('why', '')),
            str(p.get('horizon', 'Next 30 days'))
        ]
        ws_priorities.append(row_vals)
        ws_priorities.row_dimensions[row_idx].height = 20

    max_prio_row = max(len(priorities_data)+100, 500)
    dv_prio_horizon = DataValidation(type='list', formula1=horizon_range, allow_blank=True)
    ws_priorities.add_data_validation(dv_prio_horizon)
    dv_prio_horizon.add(f"F2:F{max_prio_row}")

    # -------------------------------------------------------------
    # 4. COMPANIES SHEET
    # -------------------------------------------------------------
    ws_companies = wb.create_sheet(title="Companies")
    ws_companies.append(["Company ID", "Company Name"])
    ws_companies.row_dimensions[1].height = 26
    companies_data = state.get('settings', {}).get('companies', [])
    for row_idx, c in enumerate(companies_data, start=2):
        ws_companies.append([str(c.get('id', '')), str(c.get('name', ''))])
        ws_companies.row_dimensions[row_idx].height = 20

    # -------------------------------------------------------------
    # Apply Formatting, Colors & Auto-Widths to all visible sheets
    # -------------------------------------------------------------
    for ws in [ws_actions, ws_decisions, ws_priorities, ws_companies]:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Format header row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = data_font
                cell.alignment = Alignment(vertical="center")

                # ID column style
                if cell.column == 1:
                    cell.font = id_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Color-code status cells
                if ws.title == "Actions" and cell.column == 5:  # Status column in Actions
                    if cell.value:
                        fill_style, font_style = get_status_style(str(cell.value))
                        cell.fill = fill_style
                        cell.font = font_style
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ws.title == "Decisions" and cell.column == 4:  # Status column in Decisions
                    if cell.value:
                        fill_style, font_style = get_status_style(str(cell.value))
                        cell.fill = fill_style
                        cell.font = font_style
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Set width with breathing room (min 13, max 60)
            ws.column_dimensions[col_letter].width = max(min(max_len + 5, 60), 13)

    # Remove the initial empty sheet
    if default_sheet in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(default_sheet)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_state_to_csv_zip(state: Dict[str, Any]) -> io.BytesIO:
    """Exports state as a ZIP archive containing individual CSVs."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        df_actions = pd.DataFrame(state.get('actions', []))
        zf.writestr('actions.csv', df_actions.to_csv(index=False))

        df_decisions = pd.DataFrame(state.get('decisions', []))
        zf.writestr('decisions.csv', df_decisions.to_csv(index=False))

        df_priorities = pd.DataFrame(state.get('priorities', []))
        zf.writestr('priorities.csv', df_priorities.to_csv(index=False))

    zip_buffer.seek(0)
    return zip_buffer
